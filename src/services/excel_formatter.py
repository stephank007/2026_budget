# excel_formatter.py
from __future__ import annotations

from typing import Dict, List

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment


AMOUNT_LIKE_KEYWORDS: List[str] = ["amount", "סכום"]
AMOUNT_EXACT_NAMES = {
    "amount",
    "mc_amount",
    "raw_amount",
    "סכום",
    "סכום חיוב",
    "סכום עסקה",
    "סכום עמלה",
}


def _style_header_row(ws: Worksheet) -> None:
    """
    Apply a consistent style to the header row:
    - bold font
    - centered text
    - light blue fill
    """
    if ws.max_column < 1:
        return
    
    header_fill = PatternFill(
        start_color="D9E1F2",
        end_color="D9E1F2",
        fill_type="solid",
    )
    header_font = Font(bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _format_amount_columns(ws: Worksheet) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    
    header: Dict[int, str] = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        name = str(cell.value) if cell.value is not None else ""
        header[col_idx] = name
    
    amount_cols: List[int] = []
    for col_idx, name in header.items():
        low = name.lower()
        if name in AMOUNT_EXACT_NAMES:
            amount_cols.append(col_idx)
        elif any(k in low for k in AMOUNT_LIKE_KEYWORDS):
            amount_cols.append(col_idx)
    
    for col_idx in amount_cols:
        for row_idx in range(2, ws.max_row + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            if c.value is not None and isinstance(c.value, (int, float)):
                c.number_format = "#,##0.00"


def _auto_width(ws: Worksheet, padding: int = 2) -> None:
    if ws.max_column < 1:
        return
    
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                val = str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
        ws.column_dimensions[col_letter].width = max_len + padding


def _add_header_filter_and_freeze(ws: Worksheet) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def format_workbook_default(wb: Workbook) -> None:
    """
    Apply default formatting to all worksheets:
      * header style (bold, centered, light blue)
      * detect amount columns and format numeric
      * auto-fit widths
      * auto-filter & freeze top row
    """
    for ws in wb.worksheets:
        _style_header_row(ws)
        _format_amount_columns(ws)
        _auto_width(ws)
        _add_header_filter_and_freeze(ws)
