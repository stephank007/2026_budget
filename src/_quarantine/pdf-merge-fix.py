import re
import itertools
from typing import Any, Dict, Optional
from openpyxl.styles import PatternFill


import pandas as pd

from common_paths import MERGED_TABLES_XLSX, DATA_PROCESSED, DATA_CSV, DATA_RAW_TRANSACTIONS
from process_tnuot_transactions import process_transactions

DATE_TOKEN_RE = re.compile(r"\d{2}/\d{2}/\d{2}")
DATE_FULL_RE = re.compile(r"^(\d{2})/(\d{2})/(\d{2})$")


def fix_date(s: str) -> Optional[str]:
    if not isinstance(s, str):
        return None
    s = s.strip()
    if not s:
        return None
    
    try:
        # Handle ISO-style timestamps explicitly
        if re.match(r"\d{4}-\d{2}-\d{2}", s):
            dt = pd.to_datetime(s, format="%Y-%m-%d %H:%M:%S", errors="coerce")
        else:
            dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
    except Exception:
        dt = pd.NaT
    
    if not pd.isna(dt):
        return dt.strftime("%d.%m.%y")
    
    m = DATE_FULL_RE.match(s)
    if not m:
        return s
    a, b, c = m.groups()
    return f"{c}.{b[::-1]}.{a[::-1]}"


def contains_hebrew(s: str) -> bool:
    return any("\u0590" <= ch <= "\u05FF" for ch in s)


def is_number_like(s: str) -> bool:
    s = s.strip()
    return bool(s and re.fullmatch(r"-?[0-9.,]+-?", s))


def normalize_number(s: str) -> Optional[float]:
    if not isinstance(s, str):
        return None
    s = s.strip()
    if not s:
        return None
    if s.endswith("-") and not s.startswith("-"):
        s = "-" + s[:-1]
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_transaction_cell(raw: Any) -> Dict[str, Any]:
    if pd.isna(raw):
        return {"date": None, "payee": None, "amount": None}
    
    lines = [ln.strip() for ln in str(raw).splitlines() if ln.strip()]
    if not lines:
        return {"date": None, "payee": None, "amount": None}
    
    date_line = next((ln for ln in lines if DATE_TOKEN_RE.search(ln)), None)
    date_str = fix_date(date_line) if date_line else None
    
    payee_idx = next((i for i, ln in enumerate(lines) if contains_hebrew(ln)), None)
    if payee_idx is None:
        return {"date": date_str, "payee": None, "amount": None}
    
    payee_clean = (
        lines[payee_idx].replace("(", "").replace(")", "").replace('"', "").strip()
    )
    payee = payee_clean[::-1]
    
    amount = None
    if payee_idx > 0 and is_number_like(lines[payee_idx - 1]):
        amount = normalize_number(lines[payee_idx - 1])
    
    return {"date": date_str, "payee": payee, "amount": amount}


def parse_transaction_row(row: pd.Series) -> Dict[str, Any]:
    """
    Handle both Camelot formats:

    1. Legacy format: everything is in col_1 as multi-line text.
       In this case, all other col_* are empty -> use parse_transaction_cell().
    2. Structured format: numeric columns, Hebrew payee in e.g. col_4,
       date in e.g. col_5. Here we parse directly from the columns.
    """
    
    # Columns that might contain transaction data
    col_names = [c for c in row.index if str(c).startswith("col_")]
    
    # Check if there is any non-empty value outside col_1
    other_cols = [c for c in col_names if c != "col_1"]
    has_other_data = any(
        pd.notna(row.get(c)) and str(row.get(c)).strip() for c in other_cols
    )
    
    # ---- Case 1: legacy "all in col_1" format ----
    if not has_other_data:
        return parse_transaction_cell(row.get("col_1"))
    
    # ---- Case 2: structured columns ----
    
    # Date: in your screenshot it’s in col_5 ("2025-01-05 00:00:00")
    raw_date = row.get("col_5")
    date_str = fix_date(str(raw_date)) if pd.notna(raw_date) else None
    
    # Amount: in your screenshot it's in col_3 ("2.58")
    raw_amount = row.get("col_3")
    amount = None
    if pd.notna(raw_amount):
        amount = normalize_number(str(raw_amount))
    
    # Payee: Hebrew text, in your screenshot col_4 ("קרן מכבי (י)")
    raw_payee = row.get("col_4")
    payee = None
    if isinstance(raw_payee, str) and raw_payee.strip():
        # IMPORTANT: for structured tables, the Hebrew is already in correct order
        payee_clean = (
            raw_payee.replace("(", "").replace(")", "").replace('"', "").strip()
        )
        payee = payee_clean
    
    return {"date": date_str, "payee": payee, "amount": amount}


def read_and_filter_card_csvs(
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
) -> pd.DataFrame:
    """
    Read all CSV files in data/csv that match the expected structure,
    merge them, and filter rows where 'תאריך' is within the given date range.
    """
    
    expected_cols = [
        "סטטוס",
        "תאור",
        "סכום עמלה",
        "סכום",
        "אמצעי תשלום",
        "זיכוי/חיוב",
        "מאת/ל",
        "תאריך",
    ]
    
    frames = []
    
    for csv_path in DATA_CSV.glob("*.csv"):
        try:
            df_csv = pd.read_csv(csv_path, encoding="utf-8-sig")
        except Exception as e:
            print(f"Skipping {csv_path} (read error: {e})")
            continue
        
        cols_no_unnamed = [c for c in df_csv.columns if not str(c).startswith("Unnamed")]
        
        if not set(expected_cols).issubset(set(cols_no_unnamed)):
            print(f"Skipping {csv_path} (unexpected columns)")
            continue
        
        frames.append(df_csv)
    
    if not frames:
        print("No matching CSV files found in data/csv.")
        return pd.DataFrame(columns=expected_cols)
    
    merged = pd.concat(frames, ignore_index=True)
    
    # parse dd.mm.yy
    csv_dates = pd.to_datetime(
        merged["תאריך"], format="%d.%m.%y", errors="coerce",
    )
    
    mask = (csv_dates >= start_date) & (csv_dates <= end_date)
    
    filtered = merged.loc[mask].copy()
    
    print(
        f"Loaded {len(merged)} rows from CSV files, "
        f"{len(filtered)} fall within {start_date.date()}–{end_date.date()}",
    )
    
    return filtered


def refine_payees_from_card(df1: pd.DataFrame, card_df: pd.DataFrame) -> pd.DataFrame:
    """
    Update df1['payee'] based on card_df, using 2 rules:

    1. For amounts that have a 'משיכה' row in card_df and another row
       with the same amount: use the *other* row's 'תאור' as payee.
    2. For the rest, match by amount and use card_df['תאור'] as payee.
    Matching is done on *absolute* amount.
    """
    
    if card_df.empty:
        return df1
    
    df1 = df1.copy()
    card_df = card_df.copy()
    
    # numeric amounts in card_df
    card_df["amount_num"] = card_df["סכום"].astype(str).apply(normalize_number)
    card_df = card_df.dropna(subset=["amount_num"])
    
    card_df["amount_abs"] = card_df["amount_num"].abs()
    df1["amount_abs"] = df1["amount"].abs()
    
    # ---------- Rule 1: "משיכה" special case ----------
    is_atm = card_df["תאור"].astype(str).str.contains("משיכה", na=False)
    atm_rows = card_df[is_atm]
    
    atm_map = {}  # amount_abs -> תיאור (non-משיכה)
    for amt in atm_rows["amount_abs"].unique():
        group = card_df[card_df["amount_abs"] == amt]
        non_atm = group[~group["תאור"].astype(str).str.contains("משיכה", na=False)]
        if not non_atm.empty:
            atm_map[amt] = non_atm.iloc[0]["תאור"]
    
    # apply rule 1 to df1: rows whose abs(amount) appears in atm_map
    if atm_map:
        atm_mask_df1 = df1["amount_abs"].isin(atm_map.keys())
        df1.loc[atm_mask_df1, "payee"] = df1.loc[atm_mask_df1, "amount_abs"].map(atm_map)
    
    # ---------- Rule 2: generic amount → תיאור mapping ----------
    non_atm_df = card_df[~card_df["תאור"].astype(str).str.contains("משיכה", na=False)].copy()
    
    general_map = (
        non_atm_df.dropna(subset=["amount_abs"])
        .drop_duplicates(subset=["amount_abs"])
        .set_index("amount_abs")["תאור"]
        .to_dict()
    )
    
    # Apply rule 2 to all rows that were NOT handled by rule 1
    remaining_mask = ~df1["amount_abs"].isin(atm_map.keys())
    df1.loc[remaining_mask, "payee"] = (
        df1.loc[remaining_mask, "amount_abs"].map(general_map)
        .fillna(df1.loc[remaining_mask, "payee"])
    )
    
    # clean up helper column
    df1.drop(columns=["amount_abs"], inplace=True, errors="ignore")
    
    return df1


def build_mastercard_matches(
    mastercard_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    max_len: int = 6,
) -> pd.DataFrame:
    """
    For each row in mastercard_df (one right charge),
    find a combination (1–max_len rows) in raw_df whose *absolute* 'amount'
    sum matches the absolute mastercard amount.

    Returns one row per matched raw transaction, with grouping columns.
    """
    mc = mastercard_df.copy()
    mc["amount"] = pd.to_numeric(mc["amount"], errors="coerce").round(2)
    mc = mc.dropna(subset=["amount"])
    mc["amount_abs"] = mc["amount"].abs()
    
    raw = raw_df.copy()
    
    if "amount" not in raw.columns:
        return pd.DataFrame(
            columns=[
                "group_id", "combo_size", "match_diff",
                "mc_date", "mc_payee", "mc_amount",
                "raw_date", "raw_payee", "raw_amount",
            ]
        )
    
    raw["amount"] = pd.to_numeric(raw["amount"], errors="coerce").round(2)
    raw = raw.dropna(subset=["amount"])
    raw["amount_abs"] = raw["amount"].abs()
    
    matches = []
    available = raw.copy()
    group_id = 0
    
    for _, mc_row in mc.iterrows():
        target = mc_row["amount_abs"]
        if pd.isna(target):
            continue
        
        idxs = list(available.index)
        vals = available["amount_abs"].to_dict()
        
        found_combo = None
        found_sum = None
        
        for r in range(1, max_len + 1):
            combo_found = False
            for combo in itertools.combinations(idxs, r):
                s = sum(vals[i] for i in combo)
                if abs(s - target) < 0.01:  # tolerance
                    found_combo = combo
                    found_sum = s
                    combo_found = True
                    break
            if combo_found:
                break
        
        if found_combo is None:
            continue  # not matched
        
        group_id += 1
        combo_size = len(found_combo)
        match_diff = round(found_sum - target, 2)
        
        for ridx in found_combo:
            rr = available.loc[ridx]
            matches.append(
                {
                    "group_id": group_id,
                    "combo_size": combo_size,
                    "match_diff": match_diff,
                    "mc_date": mc_row.get("date"),
                    "mc_payee": mc_row.get("payee"),
                    "mc_amount": mc_row["amount"],  # negative
                    "raw_date": rr.get("date"),
                    "raw_payee": rr.get("payee"),
                    "raw_amount": rr["amount"],     # positive
                }
            )
        
        available = available.drop(index=list(found_combo))
    
    return pd.DataFrame(matches)


def load_and_clean_raw_9846(path) -> pd.DataFrame:
    """
    Load a 9846_mm_yyyy.xlsx file that has:
      - junk rows at the top
      - a row with Hebrew column names (תאריך רכישה, שם בית עסק, סכום עסקה, ...)
      - then the real data rows.
    Return a dataframe with columns: date, payee, amount
    """
    # Read WITHOUT treating any row as header
    df = pd.read_excel(path, header=None)
    
    # Drop completely empty rows
    df = df.dropna(how="all")
    if df.empty:
        return pd.DataFrame(columns=["date", "payee", "amount"])
    
    # find the row that contains "תאריך רכישה" anywhere
    header_mask = df.apply(
        lambda row: row.astype(str).str.contains("תאריך רכישה", na=False).any(),
        axis=1,
    )
    header_idxs = df.index[header_mask]
    
    if header_idxs.empty:
        print(f"Warning: couldn't find 'תאריך רכישה' header in {path}")
        return pd.DataFrame(columns=["date", "payee", "amount"])
    
    header_idx = header_idxs[0]
    
    # This row will become the columns
    header_row = df.loc[header_idx]
    
    # Data is all rows AFTER the header row
    data = df.loc[df.index > header_idx].copy()
    data.columns = header_row
    
    # Drop columns that are entirely NaN (just in case)
    data = data.dropna(axis=1, how="all")
    
    # Rename Hebrew columns → English names expected by the matcher
    rename_map = {}
    cols = data.columns.astype(str)
    
    if "תאריך רכישה" in cols:
        rename_map["תאריך רכישה"] = "date"
    if "שם בית עסק" in cols:
        rename_map["שם בית עסק"] = "payee"
    if "סכום חיוב" in cols:
        rename_map["סכום חיוב"] = "amount"
    elif "סכום עסקה" in cols:
        rename_map["סכום עסקה"] = "amount"
    
    data = data.rename(columns=rename_map)
    
    keep_cols = [c for c in ["date", "payee", "amount"] if c in data.columns]
    data = data[keep_cols]
    
    if "payee" in data.columns:
        data["payee"] = data["payee"].astype(str).str.strip()
    
    return data


def main():
    input_path = MERGED_TABLES_XLSX
    
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")
    
    df = pd.read_excel(input_path)
    
    if "col_1" not in df.columns:
        raise SystemExit("Column 'col_1' not found in merged_tables.xlsx")
    
    # Parse each row, supporting both “all-in-col_1” and structured formats
    parsed_df = df.apply(parse_transaction_row, axis=1, result_type="expand")
    
    parsed_df["amount"] = pd.to_numeric(parsed_df["amount"], errors="coerce").round(2)
    
    blocked_keywords = [
        "מסטרקארד",
        "אמריקן",
        "דיינרס",
        # "זיכוי",
    ]
    
    pattern = "|".join(blocked_keywords)
    mask = parsed_df["payee"].str.contains(pattern, na=False)
    
    filtered_df = parsed_df[~mask].copy()
    
    # store blocked separately
    blocked_df = parsed_df[mask].copy()
    
    # ---- determine months/years present & export one file per month ----
    
    date_dt = pd.to_datetime(filtered_df["date"], format="%d.%m.%y", errors="coerce")
    
    if date_dt.notna().sum() == 0:
        raise SystemExit("No valid dates found; cannot determine mm_yy.")
    
    # keep only rows with valid dates
    valid_mask = date_dt.notna()
    filtered_df_valid = filtered_df.loc[valid_mask].copy()
    date_dt = date_dt[valid_mask]
    
    # (year, month) pairs present in the data
    months_years = sorted(
        set(zip(date_dt.dt.year.astype(int), date_dt.dt.month.astype(int)))
    )
    
    DATA_PROCESSED.mkdir(parents=True, exist_ok=True)
    
    for year, month in months_years:
        # rows for this specific month/year
        idx = (date_dt.dt.year == year) & (date_dt.dt.month == month)
        df_month = filtered_df_valid.loc[idx].copy()
        if df_month.empty:
            continue
        
        mm = f"{month:02d}"
        yy = f"20{year % 100:02d}"
        
        # dates for blocked_df
        blocked_dates = pd.to_datetime(
            blocked_df["date"], format="%d.%m.%y", errors="coerce"
        )
        blocked_month = blocked_df[
            (blocked_dates.dt.year == year) & (blocked_dates.dt.month == month)
            ].copy()
        
        # מסטרקארד rows for this month
        mastercard_month = blocked_month[
            blocked_month["payee"].astype(str).str.contains("מסטרקארד", na=False)
        ].copy()
        
        # ---- load raw 9846 file for this month (if exists) & build matches ----
        raw_9846_path = DATA_RAW_TRANSACTIONS / f"9846_{mm}_{yy}.xlsx"
        
        mastercard_matches_df = pd.DataFrame()
        if not mastercard_month.empty and raw_9846_path.exists():
            raw_df = load_and_clean_raw_9846(raw_9846_path)
            if not raw_df.empty:
                mastercard_matches_df = build_mastercard_matches(
                    mastercard_month, raw_df, max_len=6
                )
        
        # ---- build combined Mastercard sheet with status ----
        if not mastercard_matches_df.empty:
            mc_matches = mastercard_matches_df.copy()
            mc_matches["status"] = "matched"
            matched_keys = set(
                zip(
                    mc_matches["mc_date"],
                    mc_matches["mc_payee"],
                    mc_matches["mc_amount"],
                )
            )
        else:
            mc_matches = pd.DataFrame(
                columns=[
                    "group_id",
                    "combo_size",
                    "match_diff",
                    "mc_date",
                    "mc_payee",
                    "mc_amount",
                    "raw_date",
                    "raw_payee",
                    "raw_amount",
                    "status",
                ]
            )
            matched_keys = set()
        
        unmatched_rows = []
        for _, row in mastercard_month.iterrows():
            key = (row.get("date"), row.get("payee"), row.get("amount"))
            if key not in matched_keys:
                unmatched_rows.append(
                    {
                        "group_id": None,
                        "combo_size": 0,
                        "match_diff": None,
                        "mc_date": row.get("date"),
                        "mc_payee": row.get("payee"),
                        "mc_amount": row.get("amount"),
                        "raw_date": None,
                        "raw_payee": None,
                        "raw_amount": None,
                        "status": "unmatched",
                    }
                )
        
        if unmatched_rows:
            unmatched_df = pd.DataFrame(unmatched_rows)
            mc_sheet_df = pd.concat([mc_matches, unmatched_df], ignore_index=True)
        else:
            mc_sheet_df = mc_matches.copy()
        
        total_mc = len(mastercard_month)
        matched_mc = len(matched_keys)
        unmatched_mc = total_mc - matched_mc
        
        # ---- date range for CSV filter for this month ----
        month_dates = pd.to_datetime(df_month["date"], format="%d.%m.%y", errors="coerce")
        month_dates = month_dates.dropna()
        start_date = month_dates.min().normalize()
        end_date = month_dates.max().normalize()
        
        # ---- read card CSVs & keep only 'בוצע' for this month range ----
        card_df = read_and_filter_card_csvs(start_date, end_date)
        card_df = card_df[card_df["סטטוס"] == "בוצע"].copy()
        
        # ---- process PDF transactions for this month only ----
        output_path = DATA_PROCESSED / f"9016_{mm}_{yy}.xlsx"
        
        df1, df2 = process_transactions(df=df_month)
        df1 = df1.dropna(subset=["amount"])
        df1.sort_values("amount", inplace=True)
        
        # ---- refine df1 payees based on card_df rules ----
        if not card_df.empty:
            df1 = refine_payees_from_card(df1, card_df)
        
        # optional: for debugging – see date range used
        dates = pd.to_datetime(df1["date"], format="%d.%m.%y", errors="coerce").dropna()
        if not dates.empty:
            print(
                f"{mm}.{yy} range:",
                dates.min().strftime("%Y-%m"),
                dates.max().strftime("%Y-%m"),
            )
        
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # 1) Write all sheets first
            df1.to_excel(writer, sheet_name="Parsed", index=False)
            df2.to_excel(writer, sheet_name="Income", index=False)
            
            # ---- Mastercard summary sheet ----
            summary_df = pd.DataFrame(
                [
                    {
                        "month"                    : f"{mm}.{yy}",
                        "total_mastercard_rows"    : total_mc,
                        "matched_mastercard_rows"  : matched_mc,
                        "unmatched_mastercard_rows": unmatched_mc,
                    }
                ]
            )
            summary_df.to_excel(writer, sheet_name="MC_Summary", index=False)
            
            # ---- blocked keyword sheets per MONTH ----
            for kw in blocked_keywords:
                kw_month_df = blocked_month[
                    blocked_month["payee"].astype(str).str.contains(kw, na=False)
                ].copy()
                
                if kw == "מסטרקארד":
                    # always use combined matches+unmatched with status
                    sheet_df = mc_sheet_df
                else:
                    sheet_df = kw_month_df
                
                if sheet_df is not None and not sheet_df.empty:
                    sheet_name = kw[:28]  # Excel sheet name length safety
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 2) NOW do formatting, after all sheets exist
            wb = writer.book
            
            # 2.1 All sheets: format "amount" columns as #,##0.00
            amount_like_keywords = ["amount", "סכום"]
            amount_exact_names = {
                "amount",
                "mc_amount",
                "raw_amount",
                "סכום",
                "סכום חיוב",
                "סכום עסקה",
                "סכום עמלה",
            }
            
            for ws in wb.worksheets:
                # assume first row is header
                header = {}
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    name = str(cell.value) if cell.value is not None else ""
                    header[col_idx] = name
                
                # decide which columns are "amount" columns
                amount_cols = []
                for col_idx, name in header.items():
                    low = name.lower()
                    if name in amount_exact_names:
                        amount_cols.append(col_idx)
                    elif any(k in low for k in amount_like_keywords):
                        amount_cols.append(col_idx)
                
                # apply number format to all data rows for these columns
                for col_idx in amount_cols:
                    for row_idx in range(2, ws.max_row + 1):
                        c = ws.cell(row=row_idx, column=col_idx)
                        if c.value is not None and isinstance(c.value, (int, float)):
                            c.number_format = "#,##0.00"
            
            # 2.2 Parsed sheet: auto column width
            if "Parsed" in wb.sheetnames:
                ws_parsed = wb["Parsed"]
                for col in ws_parsed.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value is not None:
                            val = str(cell.value)
                            if len(val) > max_len:
                                max_len = len(val)
                    # small padding
                    ws_parsed.column_dimensions[col_letter].width = max_len + 2
                    
            if "Income" in wb.sheetnames:
                ws_parsed = wb["Income"]
                for col in ws_parsed.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value is not None:
                            val = str(cell.value)
                            if len(val) > max_len:
                                max_len = len(val)
                    # small padding
                    ws_parsed.column_dimensions[col_letter].width = max_len + 2

            # 2.3 מסטרקארד sheet: auto width + color groups by mc_amount
            mc_sheet_name = "מסטרקארד"
            if mc_sheet_name in wb.sheetnames:
                ws_mc = wb[mc_sheet_name]
                
                # auto column width
                for col in ws_mc.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value is not None:
                            val = str(cell.value)
                            if len(val) > max_len:
                                max_len = len(val)
                    ws_mc.column_dimensions[col_letter].width = max_len + 2
                
                # find mc_amount column index
                mc_amount_col_idx = None
                for col_idx in range(1, ws_mc.max_column + 1):
                    hdr_val = ws_mc.cell(row=1, column=col_idx).value
                    if hdr_val == "mc_amount":
                        mc_amount_col_idx = col_idx
                        break
                
                if mc_amount_col_idx is not None:
                    # define a palette of light fills
                    fills = [
                        PatternFill(start_color="FFFDE9D9", end_color="FFFDE9D9", fill_type="solid"),
                        PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"),
                        PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid"),
                        PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid"),
                        PatternFill(start_color="FFF6E5F7", end_color="FFF6E5F7", fill_type="solid"),
                        PatternFill(start_color="FFE4DFEC", end_color="FFE4DFEC", fill_type="solid"),
                    ]
                    
                    amount_to_fill = {}
                    fill_idx = 0
                    
                    # start from row 2 (skip header)
                    for row_idx in range(2, ws_mc.max_row + 1):
                        cell = ws_mc.cell(row=row_idx, column=mc_amount_col_idx)
                        val = cell.value
                        if val is None:
                            continue
                        
                        # use value as key; same mc_amount => same color
                        key = float(val) if isinstance(val, (int, float)) else str(val)
                        
                        if key not in amount_to_fill:
                            amount_to_fill[key] = fills[fill_idx % len(fills)]
                            fill_idx += 1
                        
                        row_fill = amount_to_fill[key]
                        
                        # color the entire row for this mc_amount
                        for col_idx in range(1, ws_mc.max_column + 1):
                            ws_mc.cell(row=row_idx, column=col_idx).fill = row_fill
        
        print(f"Done. Wrote {len(df_month)} rows to: {output_path.resolve()}")
        
if __name__ == "__main__":
    main()
