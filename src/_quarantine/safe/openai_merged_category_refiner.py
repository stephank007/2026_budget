from __future__ import annotations

import json
from typing import Dict, Optional, Tuple, Any

import pandas as pd
from openpyxl import load_workbook
from openai import OpenAI

from common_paths import DATA_PROCESSED, CONFIG_DIR
from services.payee_service import load_payee_resources

# OPTIONAL: fuzzy matching for payee names
try:
    from rapidfuzz import fuzz, process as rf_process
except ImportError:
    fuzz = None
    rf_process = None


# -----------------------------------------------------------------------------
# OpenAI client
# -----------------------------------------------------------------------------
client = OpenAI()


# -----------------------------------------------------------------------------
# Garbage detection
# -----------------------------------------------------------------------------
GARBAGE_PREFIXES: Tuple[str, ...] = (
    "העב.",
    "משיכת",
    "זיכוי",
)

GARBAGE_KEYWORDS: Tuple[str, ...] = (
    "באינטרנט",
    "מידית",
    "הוראת קבע",
)


def looks_like_garbage(payee: str) -> bool:
    """
    Return True if the payee looks like a generic/non-merchant string that
    shouldn't be classified (e.g. transfer, standing order, etc.).
    """
    if not isinstance(payee, str):
        return True
    
    s = payee.strip()
    if not s:
        return True
    
    for p in GARBAGE_PREFIXES:
        if s.startswith(p):
            return True
    
    for kw in GARBAGE_KEYWORDS:
        if kw in s:
            return True
    
    return False


# -----------------------------------------------------------------------------
# Config loading (heuristics live in config/category_heuristics.json)
# -----------------------------------------------------------------------------
def load_category_heuristics() -> Dict[str, Any]:
    """
    Load category heuristics from:

        <project_root>/config/category_heuristics.json

    Expected structure example:

    {
      "restaurant": {
        "category_name": "מסעדות ובתי קפה",
        "keywords": ["מסעדה", "קפה", "פיצה"],
        "name_hints": ["גרקו"]
      }
    }

    If the file is missing or invalid, we return an empty dict,
    which simply disables heuristics (safe-by-default).
    """
    
    config_path = CONFIG_DIR / "category_heuristics.json"
    
    try:
        with config_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        
        if not isinstance(data, dict):
            print(f"[refiner] config {config_path} is not a JSON object — ignoring")
            return {}
        
        return data
    
    except FileNotFoundError:
        print(f"[refiner] config file not found: {config_path}")
    except Exception as e:
        print(f"[refiner] error reading config {config_path}: {e}")
    
    # Safe fallback — heuristics disabled
    return {}


CATEGORY_HEURISTICS = load_category_heuristics()

# Convenience accessors (safe defaults for restaurant heuristics)
_RESTAURANT_CFG = CATEGORY_HEURISTICS.get("restaurant", {})

RESTAURANT_CATEGORY_NAME: str = _RESTAURANT_CFG.get("category_name", "").strip()

RESTAURANT_KEYWORDS: Tuple[str, ...] = tuple(
    k for k in _RESTAURANT_CFG.get("keywords", []) if isinstance(k, str)
)

RESTAURANT_NAME_HINTS: Tuple[str, ...] = tuple(
    k for k in _RESTAURANT_CFG.get("name_hints", []) if isinstance(k, str)
)


# -----------------------------------------------------------------------------
# Lookup / rules
# -----------------------------------------------------------------------------
def build_lookup_dicts():
    """
    Reuse existing payee resources already used in bank_fixer:
      - rules_df, lookup_df = load_payee_resources()
    """
    rules_df, lookup_df = load_payee_resources()
    
    payee_to_cat = (
        lookup_df
        .dropna(subset=["payee", "category"])
        .drop_duplicates("payee")
        .set_index("payee")["category"]
        .to_dict()
    )
    
    rules = (
        rules_df
        .dropna(subset=["match_string", "normalized_payee"])
        [["match_string", "normalized_payee"]]
        .to_dict(orient="records")
    )
    
    allowed_categories = sorted(lookup_df["category"].dropna().unique().tolist())
    
    return payee_to_cat, rules, allowed_categories, lookup_df


def rule_based_guess(
    payee: str,
    rules,
    payee_to_cat: Dict[str, str],
) -> Optional[str]:
    """
    1. Exact payee → category from lookup.
    2. Substring rule (match_string) → normalized_payee → lookup.
    """
    if not isinstance(payee, str):
        return None
    payee = payee.strip()
    if not payee:
        return None
    
    # 1) exact lookup
    if payee in payee_to_cat:
        return payee_to_cat[payee]
    
    # 2) substring rule → normalized_payee → lookup
    for rule in rules:
        m = rule.get("match_string")
        if m and m in payee:
            normalized = rule.get("normalized_payee")
            if normalized in payee_to_cat:
                return payee_to_cat[normalized]
    
    return None


def fuzzy_guess(
    payee: str,
    payee_to_cat: Dict[str, str],
    min_score: int = 90,
) -> Optional[str]:
    """
    Fuzzy match payee name against known payees and return the matched category
    if the similarity score is high enough.
    """
    if fuzz is None or rf_process is None:
        return None
    
    choices = list(payee_to_cat.keys())
    if not choices:
        return None
    
    result = rf_process.extractOne(payee, choices, scorer=fuzz.token_set_ratio)
    if not result:
        return None
    
    best_match, score, _ = result
    if score >= min_score:
        return payee_to_cat[best_match]
    
    return None


# -----------------------------------------------------------------------------
# Heuristic keyword-based classification (config-driven)
# -----------------------------------------------------------------------------
def heuristic_keyword_guess(payee: str, allowed_categories) -> Optional[str]:
    """
    Cheap deterministic heuristic based on keywords / known restaurant names.

    Currently only handles 'restaurant' heuristics (from JSON config), but can
    be extended later.

    If the payee looks like a restaurant/cafe and RESTAURANT_CATEGORY_NAME is
    in allowed_categories, return that category. Otherwise return None.
    """
    if not isinstance(payee, str):
        return None
    
    s = payee.strip()
    if not s or not allowed_categories:
        return None
    
    if not RESTAURANT_CATEGORY_NAME:
        return None
    
    if RESTAURANT_CATEGORY_NAME not in allowed_categories:
        return None
    
    haystack = s.replace("-", " ")
    
    # Keyword-based hint (e.g. מסעדה, קפה, כשר, כשרה...)
    for kw in RESTAURANT_KEYWORDS:
        if kw and kw in haystack:
            return RESTAURANT_CATEGORY_NAME
    
    # Brand/name hints (e.g. גרקו)
    for name_hint in RESTAURANT_NAME_HINTS:
        if name_hint and name_hint in haystack:
            return RESTAURANT_CATEGORY_NAME
    
    return None


# -----------------------------------------------------------------------------
# LLM-based classification (with "better unknown than wrong" behavior)
# -----------------------------------------------------------------------------
def llm_guess_category(payee: str, allowed_categories) -> Optional[str]:
    """
    Use OpenAI to guess a category from allowed_categories for a payee.
    Returns category string or None.

    IMPORTANT:
      - Prompt explicitly allows / encourages "לא מזוהה" when not confident.
      - We treat 'לא מזוהה' / 'unknown' as None
        (better unknown than wrong classification).
    """
    if not payee or not allowed_categories:
        return None
    
    categories_list = "\n".join(f"- {c}" for c in allowed_categories)
    
    prompt = f"""
אתה מערכת שמסווגת עסקאות בנק לקטגוריות הוצאות.

שם בית העסק: "{payee}"

רשימת הקטגוריות המותרות:
{categories_list}

הוראות חשובות:
- אם שם העסק הוא מסעדה, בית קפה, רשת אוכל, בר, קונדיטוריה או עסק מזון אחר –
  בחר תמיד את הקטגוריה "{RESTAURANT_CATEGORY_NAME}" אם היא קיימת ברשימה.
- מילים כמו "כשר" או "כשרה" מתייחסות בדרך כלל למקומות אוכל (מסעדות, בתי קפה) ולא לספורט.
- אם אינך בטוח או שאין התאמה טובה, החזר בדיוק: "לא מזוהה".

בחר אך ורק קטגוריה אחת מהרשימה, או "לא מזוהה".
החזר רק את שם הקטגוריה (או "לא מזוהה"), בלי טקסט נוסף.
"""
    
    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You classify bank transactions into expense categories. "
                        "If you are not confident, respond with the Hebrew phrase 'לא מזוהה' exactly."
                    ),
                },
                {
                    "role": "user",
                    "content": prompt,
                },
            ],
            temperature=0,
            max_tokens=20,
        )
    except Exception as e:
        print(f"[refiner] LLM error for payee '{payee}': {e}")
        return None
    
    raw = (resp.choices[0].message.content or "").strip()
    raw = raw.replace("קטגוריה:", "").strip().strip('"\'')
    if not raw:
        return None
    
    # If model explicitly says it doesn't know → treat as unknown
    unknown_markers = {
        "לא מזוהה",
        "לא ידוע",
        "לא בטוח",
        "unknown",
        "uncertain",
        "UNK",
    }
    if raw in unknown_markers:
        return None
    
    # exact match first
    if raw in allowed_categories:
        return raw
    
    # then fuzzy against allowed categories (conservative)
    if fuzz and rf_process:
        result = rf_process.extractOne(raw, allowed_categories, scorer=fuzz.ratio)
        if result:
            match, score, _ = result
            # conservative threshold: if the LLM output is far from any real category,
            # we prefer "unknown" over a bad guess.
            if score >= 80:
                return match
    
    # If we couldn't confidently map the answer to a known category → None
    return None


# -----------------------------------------------------------------------------
# Excel helpers
# -----------------------------------------------------------------------------
def _find_header_col(ws, target_col_name: str) -> int:
    """Find which Excel column (1-based index) has name target_col_name in row 1."""
    for cell in ws[1]:
        if isinstance(cell.value, str) and cell.value.strip() == target_col_name:
            return cell.column
    raise ValueError(f"Column '{target_col_name}' not found in header row of sheet '{ws.title}'")


def _get_or_create_header_col(ws, target_col_name: str) -> int:
    """
    Return column index for target_col_name in header row,
    creating a new column at the end if it doesn't exist.
    """
    for cell in ws[1]:
        if isinstance(cell.value, str) and cell.value.strip() == target_col_name:
            return cell.column
    new_col_idx = ws.max_column + 1
    ws.cell(row=1, column=new_col_idx).value = target_col_name
    return new_col_idx


# -----------------------------------------------------------------------------
# Main entry: refine_merged_categories
# -----------------------------------------------------------------------------
def refine_merged_categories(
    mm: str,
    yy: str,
    sheet_name: str = "Parsed",
    payee_col: str = "bank_normal",
    category_col: str = "category",
) -> None:
    """
    Open DATA_PROCESSED / f"9016_{mm}_{yy}.xlsx", take the given sheet,
    and auto-fill rows where category is False/'לא מזוהה'/NaN using:

      - existing payee rules/lookup
      - fuzzy matching (optional)
      - keyword heuristics (from JSON config)
      - LLM (optional, conservative)

    Additionally:
      - Adds/uses a 'category_source' column.
      - Writes 'rule' / 'fuzzy' / 'heuristic' / 'llm' to indicate source.
    """
    xlsx_path = DATA_PROCESSED / f"9016_{mm}_{yy}.xlsx"
    
    if not xlsx_path.exists():
        print(f"[refiner] file not found: {xlsx_path}")
        return
    
    print(f"[refiner] loading (pandas view): {xlsx_path} (sheet='{sheet_name}')")
    try:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
    except ValueError:
        print(f"[refiner] sheet '{sheet_name}' not found in {xlsx_path}")
        return
    
    if payee_col not in df.columns or category_col not in df.columns:
        print(f"[refiner] expected columns '{payee_col}' & '{category_col}' not found")
        return
    
    # build lookup dicts using existing service
    payee_to_cat, rules, allowed_categories, _ = build_lookup_dicts()
    
    cat = df[category_col]
    unknown_mask = (
            cat.isna()
            | (cat == False)  # boolean False
            | (cat.astype(str).str.strip() == "לא מזוהה")
    )
    
    unknown_df = df.loc[unknown_mask].copy()
    print(f"[refiner] rows with unknown category: {len(unknown_df)}")
    
    unique_payees = (
        unknown_df[payee_col]
        .dropna()
        .astype(str)
        .str.strip()
        .unique()
        .tolist()
    )
    
    payee_to_new_cat: Dict[str, str] = {}
    payee_to_method: Dict[str, str] = {}  # "rule", "fuzzy", "heuristic", "llm"
    
    # --- Guess categories per unique payee ---
    for p in unique_payees:
        if looks_like_garbage(p):
            continue
        
        # 1) rule-based
        guess = rule_based_guess(p, rules, payee_to_cat)
        if guess:
            payee_to_new_cat[p] = guess
            payee_to_method[p] = "rule"
            continue
        
        # 2) fuzzy against known payees
        guess = fuzzy_guess(p, payee_to_cat)
        if guess:
            payee_to_new_cat[p] = guess
            payee_to_method[p] = "fuzzy"
            continue
        
        # 3) cheap heuristics (keywords / known restaurant names from config)
        guess = heuristic_keyword_guess(p, allowed_categories)
        if guess:
            payee_to_new_cat[p] = guess
            payee_to_method[p] = "heuristic"
            continue
        
        # 4) LLM (with "לא מזוהה" / None fallback)
        guess = llm_guess_category(p, allowed_categories)
        if guess:
            print(f'using openai LLM for: {p}')
            payee_to_new_cat[p] = guess
            payee_to_method[p] = "llm"
    
    print(f"[refiner] guessed categories for {len(payee_to_new_cat)} unique payees")
    
    if not payee_to_new_cat:
        print("[refiner] nothing to update")
        return
    
    # --- Now: update only the category cells via openpyxl (keep formatting) ---
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        print(f"[refiner] sheet '{sheet_name}' not found in workbook (openpyxl)")
        return
    
    ws = wb[sheet_name]
    
    try:
        category_col_idx = _find_header_col(ws, category_col)
        payee_col_idx = _find_header_col(ws, payee_col)
    except ValueError as e:
        print(f"[refiner] {e}")
        return
    
    # ensure we have a column to mark decision source
    source_col_name = "category_source"
    source_col_idx = _get_or_create_header_col(ws, source_col_name)
    
    # df index 0 -> Excel row 2 (header in row 1)
    first_data_row = 2
    
    for df_idx in df.index[unknown_mask]:
        payee_val = df.at[df_idx, payee_col]
        if not isinstance(payee_val, str):
            payee_val = str(payee_val)
        payee_val = payee_val.strip()
        
        new_cat = payee_to_new_cat.get(payee_val)
        if not new_cat:
            continue
        
        excel_row = first_data_row + df_idx  # header offset
        
        # update category
        cat_cell = ws.cell(row=excel_row, column=category_col_idx)
        cat_cell.value = new_cat  # style remains as-is
        
        # mark source if known
        method = payee_to_method.get(payee_val)
        if method:
            src_cell = ws.cell(row=excel_row, column=source_col_idx)
            src_cell.value = method
    
    wb.save(xlsx_path)
    print(f"[refiner] updated '{sheet_name}' in {xlsx_path}")
