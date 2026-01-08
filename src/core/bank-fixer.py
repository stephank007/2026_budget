import itertools
import json
import re
from typing import Any, Iterable, Optional, Tuple

from openpyxl.styles import PatternFill
import pandas as pd

from bank_normalize_transactions import process_transactions
from openai_merged_category_refiner import refine_merged_categories

from common_paths import (
    TNUOT_9016,
    DATA_PROCESSED,
    DATA_RAW_TRANSACTIONS,
    CONFIG_JSON_PATH,
)

from excel_formatter import format_workbook_default
from services.payee_service import (
    load_payee_resources,
    apply_payee_rules_and_categories,
)
from utils import (
    fix_date,
    contains_hebrew,
    normalize_number,
)

# ---------------------------------------------------------------------------
# Constants & regexes
# ---------------------------------------------------------------------------

DATE_TOKEN_RE = re.compile(r"\d{2}/\d{2}/\d{2}")
DATE_FULL_RE = re.compile(r"^(\d{2})/(\d{2})/(\d{2})$")

# Default keywords for "blocked" payees (credit cards, etc.)
DEFAULT_BLOCKED_PAYEE_KEYWORDS = [
    "מסטרקארד",
    "אמריקן",
    "דיינרס",
    "רכישה",
]

# Keywords we create separate sheets for in the workbook
CARD_SUMMARY_SHEET_KEYWORDS = [
    "מסטרקארד",
    "אמריקן",
    "דיינרס",
]

# Keywords that define Mastercard rows within blocked transactions
MASTERCARD_IDENTIFIERS = [
    "מסטרקארד",
    "רכישה",
]

# How many months back to look when matching Mastercard charges to raw 9846 data
DEFAULT_MONTHS_BACK = 2

# Maximum number of raw 9846 rows to combine when searching for a sum match
DEFAULT_MAX_COMBO_LEN = 6

# Excel fill styles used to color groups of matching Mastercard rows
MASTERCARD_ROW_FILLS: Iterable[PatternFill] = [
    PatternFill(start_color="FFFDE9D9", end_color="FFFDE9D9", fill_type="solid"),
    PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"),
    PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid"),
    PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid"),
    PatternFill(start_color="FFF6E5F7", end_color="FFF6E5F7", fill_type="solid"),
    PatternFill(start_color="FFE4DFEC", end_color="FFE4DFEC", fill_type="solid"),
]


# ---------------------------------------------------------------------------
# Config helpers
# ---------------------------------------------------------------------------

def load_blocked_payee_keywords(
    config_path=CONFIG_JSON_PATH,
    default: Iterable[str] = DEFAULT_BLOCKED_PAYEE_KEYWORDS,
) -> list[str]:
    """
    Load a list of blocked payee keywords from config/config.json
    under the key "blocked_payee_keywords".

    If the config file or key is missing, falls back to DEFAULT_BLOCKED_PAYEE_KEYWORDS.
    """
    try:
        with config_path.open(encoding="utf-8") as f:
            config = json.load(f)
    except FileNotFoundError:
        return list(default)
    except json.JSONDecodeError:
        # If config is malformed, don't break the pipeline – just use defaults
        return list(default)
    
    blocked = config.get("blocked_payee_keywords")
    if isinstance(blocked, list) and blocked:
        return [str(x) for x in blocked]
    
    return list(default)


# ---------------------------------------------------------------------------
# Bank XLS readers / cleaners
# ---------------------------------------------------------------------------

def load_and_clean_9016_xls(path) -> pd.DataFrame:
    """
    Load a 9016 account-statement XLSX that looks like:
      - Metadata rows at the top
      - A row with Hebrew column names: תאריך, תאריך ערך, סוג תנועה, זכות, חובה, ...
      - Then the real data rows per transaction.

    Returns a dataframe with columns: date, payee, amount
      * date   – normalized to dd.mm.yy (using fix_date)
      * payee  – taken from 'סוג תנועה' (or similar)
      * amount – credit minus debit (זכות - חובה), using normalize_number
    """
    df = pd.read_excel(path, header=None)
    df = df.dropna(how="all")
    if df.empty:
        return pd.DataFrame(columns=["date", "payee", "amount"])
    
    # ---- Find the TRUE header row: has both "תאריך" and "סוג תנועה" ----
    def row_has_header_markers(row: pd.Series) -> bool:
        s = row.astype(str).copy()
        has_date = s.str.contains("תאריך", na=False).any()
        has_type = s.str.contains("סוג תנועה", na=False).any()
        return has_date and has_type
    
    header_mask = df.apply(row_has_header_markers, axis=1)
    header_idxs = df.index[header_mask]
    
    if header_idxs.empty:
        print(f"Warning: couldn't find header row with 'תאריך' + 'סוג תנועה' in {path}")
        return pd.DataFrame(columns=["date", "payee", "amount"])
    
    header_idx = header_idxs[0]
    header_row = df.loc[header_idx]
    
    # Data is all rows AFTER the header row
    data = df.loc[df.index > header_idx].copy()
    data.columns = header_row
    
    # Drop completely empty rows (bottom junk)
    data = data.dropna(how="all")
    if data.empty:
        return pd.DataFrame(columns=["date", "payee", "amount"])
    
    # ----- Detect relevant columns by HEADER TEXT -----
    cols = data.columns.astype(str)
    
    # Prefer a header that is exactly "תאריך"; fallback to "תאריך" substring.
    date_col = next((c for c in cols if str(c).strip() == "תאריך"), None)
    if date_col is None:
        date_col = next((c for c in cols if "תאריך" in str(c)), None)
    
    # Debit / credit
    debit_col = next((c for c in cols if "חובה" in str(c)), None)
    credit_col = next((c for c in cols if "זכות" in str(c)), None)
    
    # Description / payee – we really want "סוג תנועה" here
    desc_col = None
    desc_keywords = [
        "סוג תנועה",
        "תיאור",
        "תאור",
        "פרטים",
        "תיאור פעולה",
        "תיאור עסקה",
    ]
    for kw in desc_keywords:
        desc_col = next((c for c in cols if kw in str(c)), None)
        if desc_col:
            break
    
    # Fallback: pick some other non-numeric-looking column
    if desc_col is None:
        other_cols = [
            c for c in cols
            if c not in {date_col, debit_col, credit_col}
        ]
        if other_cols:
            desc_col = other_cols[0]
    
    rename_map = {}
    if date_col:
        rename_map[date_col] = "date"
    if desc_col:
        rename_map[desc_col] = "payee"
    
    data = data.rename(columns=rename_map)
    
    # ----- Build numeric amount: credit - debit using normalize_number -----
    if debit_col and debit_col in data.columns:
        debit = data[debit_col].astype(str).apply(normalize_number).fillna(0.0)
    else:
        debit = pd.Series(0.0, index=data.index)
    
    if credit_col and credit_col in data.columns:
        credit = data[credit_col].astype(str).apply(normalize_number).fillna(0.0)
    else:
        credit = pd.Series(0.0, index=data.index)
    
    data["amount"] = (credit - debit).astype(float)
    
    # ----- Normalize date and clean payee -----
    if "date" in data.columns:
        data["date"] = data["date"].apply(
            lambda x: fix_date(str(x)) if pd.notna(x) else None,
        )
    
    # Ensure a payee column always exists
    if "payee" not in data.columns:
        data["payee"] = ""
    else:
        data["payee"] = (
            data["payee"]
            .astype(str)
            .str.replace("(", "", regex=False)
            .str.replace(")", "", regex=False)
            .str.replace('"', "", regex=False)
            .str.strip()
        )
    
    # Drop header-like rows that slipped into the data
    def is_real_date(val: Any) -> bool:
        if val is None:
            return False
        s = str(val).strip()
        if not s:
            return False
        parsed = fix_date(s)
        if parsed is None:
            return False
        if contains_hebrew(parsed):
            return False
        return True
    
    mask_real = data["date"].apply(is_real_date)
    data = data[mask_real].copy()
    
    return data[["date", "payee", "amount"]].copy()


def load_and_clean_raw_9846(path) -> pd.DataFrame:
    """
    Load a 9846_mm_yyyy.xlsx file that has:
      - junk rows at the top
      - a row with Hebrew column names (תאריך רכישה, שם בית עסק, סכום עסקה, ...)
      - then the real data rows.

    Returns a dataframe with columns: date, payee, amount
    """
    df = pd.read_excel(path, header=None)
    df = df.dropna(how="all")
    if df.empty:
        return pd.DataFrame(columns=["date", "payee", "amount"])
    
    header_mask = df.apply(
        lambda row: row.astype(str).str.contains("תאריך רכישה", na=False).any(),
        axis=1,
    )
    header_idxs = df.index[header_mask]
    
    if header_idxs.empty:
        print(f"Warning: couldn't find 'תאריך רכישה' header in {path}")
        return pd.DataFrame(columns=["date", "payee", "amount"])
    
    header_idx = header_idxs[0]
    header_row = df.loc[header_idx]
    
    data = df.loc[df.index > header_idx].copy()
    data.columns = header_row
    data = data.dropna(axis=1, how="all")
    
    rename_map = {}
    cols = data.columns.astype(str)
    
    if "תאריך רכישה" in cols:
        rename_map["תאריך רכישה"] = "date"
    if "שם בית עסק" in cols:
        rename_map["שם בית עסק"] = "payee"
    if "סכום חיוב" in cols:
        rename_map["סכום חיוב"] = "amount"
    elif "סכום עסקה" in cols:
        rename_map["סכום עסקה"] = "amount"
    
    data = data.rename(columns=rename_map)
    
    keep_cols = [c for c in ["date", "payee", "amount"] if c in data.columns]
    data = data[keep_cols]
    
    if "payee" in data.columns:
        data["payee"] = data["payee"].astype(str).str.strip()
    
    return data


def load_raw_9846_window(
    mm: str,
    yy: str,
    months_back: int = DEFAULT_MONTHS_BACK,
) -> pd.DataFrame:
    """
    Load 9846_xx_yyyy.xlsx files for the current (mm, yy) and up to `months_back`
    previous months, correctly handling year wrap.

    Returns a single concatenated dataframe with columns: date, payee, amount.
    """
    base_year = int(yy)
    base_month = int(mm)
    
    dfs = []
    
    for offset in range(0, months_back + 1):
        y = base_year
        m = base_month - offset
        while m <= 0:
            m += 12
            y -= 1
        
        mm_str = f"{m:02d}"
        yy_str = f"{y:04d}"
        
        path = DATA_RAW_TRANSACTIONS / f"9846_{mm_str}_{yy_str}.xlsx"
        if path.exists():
            df = load_and_clean_raw_9846(path)
            if not df.empty:
                df["_source_file"] = path.name
                dfs.append(df)
    
    if not dfs:
        return pd.DataFrame(columns=["date", "payee", "amount", "_source_file"])
    
    return pd.concat(dfs, ignore_index=True)


# ---------------------------------------------------------------------------
# Mastercard matching & reporting
# ---------------------------------------------------------------------------

def build_mastercard_matches(
    mastercard_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    max_len: int = DEFAULT_MAX_COMBO_LEN,
) -> pd.DataFrame:
    """
    For each row in mastercard_df (one מסטרקארד charge),
    find a combination (1–max_len rows) in raw_df whose *absolute* 'amount'
    sum matches the absolute mastercard amount.

    Returns one row per matched raw transaction, with grouping columns.
    """
    mc = mastercard_df.copy()
    mc["amount"] = pd.to_numeric(mc["amount"], errors="coerce").round(2)
    mc = mc.dropna(subset=["amount"])
    mc["amount_abs"] = mc["amount"].abs()
    
    raw = raw_df.copy()
    if "amount" not in raw.columns:
        return pd.DataFrame(
            columns=[
                "group_id",
                "combo_size",
                "match_diff",
                "mc_date",
                "mc_payee",
                "mc_amount",
                "raw_date",
                "raw_payee",
                "raw_amount",
            ],
        )
    
    raw["amount"] = pd.to_numeric(raw["amount"], errors="coerce").round(2)
    raw = raw.dropna(subset=["amount"])
    raw["amount_abs"] = raw["amount"].abs()
    
    matches = []
    available = raw.copy()
    group_id = 0
    
    for _, mc_row in mc.iterrows():
        target = mc_row["amount_abs"]
        if pd.isna(target):
            continue
        
        idxs = list(available.index)
        vals = available["amount_abs"].to_dict()
        
        found_combo = None
        found_sum = None
        
        for r in range(1, max_len + 1):
            combo_found = False
            for combo in itertools.combinations(idxs, r):
                s = sum(vals[i] for i in combo)
                if abs(s - target) < 0.01:  # tolerance
                    found_combo = combo
                    found_sum = s
                    combo_found = True
                    break
            if combo_found:
                break
        
        if found_combo is None:
            continue
        
        group_id += 1
        combo_size = len(found_combo)
        match_diff = round(found_sum - target, 2)
        
        for ridx in found_combo:
            rr = available.loc[ridx]
            matches.append(
                {
                    "group_id"  : group_id,
                    "combo_size": combo_size,
                    "match_diff": match_diff,
                    "mc_date"   : mc_row.get("date"),
                    "mc_payee"  : mc_row.get("payee"),
                    "mc_amount" : mc_row["amount"],
                    "raw_date"  : rr.get("date"),
                    "raw_payee" : rr.get("payee"),
                    "raw_amount": rr["amount"],
                },
            )
        
        available = available.drop(index=list(found_combo))
    
    return pd.DataFrame(matches)


def create_mastercard_report(
    mastercard_month: pd.DataFrame,
    mm: str,
    yy: str,
) -> Tuple[pd.DataFrame, int, int]:
    """
    Build a Mastercard matching sheet + summary numbers for a given month.

    Returns:
        mc_sheet_df: dataframe for the "מסטרקארד" sheet
        total_mc: number of Mastercard rows in this month
        unmatched_mc: number of Mastercard rows that could not be matched
    """
    total_mc = len(mastercard_month)
    mastercard_matches_df = pd.DataFrame()
    
    if not mastercard_month.empty:
        raw_df = load_raw_9846_window(mm, yy, months_back=DEFAULT_MONTHS_BACK)
        if not raw_df.empty:
            mastercard_matches_df = build_mastercard_matches(
                mastercard_month,
                raw_df,
                max_len=DEFAULT_MAX_COMBO_LEN,
            )
    
    if not mastercard_matches_df.empty:
        mc_matches = mastercard_matches_df.copy()
        mc_matches["status"] = "matched"
        matched_keys = set(
            zip(
                mc_matches["mc_date"],
                mc_matches["mc_payee"],
                mc_matches["mc_amount"],
            ),
        )
    else:
        mc_matches = pd.DataFrame(
            columns=[
                "group_id",
                "combo_size",
                "match_diff",
                "mc_date",
                "mc_payee",
                "mc_amount",
                "raw_date",
                "raw_payee",
                "raw_amount",
                "status",
            ],
        )
        matched_keys = set()
    
    unmatched_rows = []
    for _, row in mastercard_month.iterrows():
        key = (row.get("date"), row.get("payee"), row.get("amount"))
        if key not in matched_keys:
            unmatched_rows.append(
                {
                    "group_id"  : None,
                    "combo_size": 0,
                    "match_diff": None,
                    "mc_date"   : row.get("date"),
                    "mc_payee"  : row.get("payee"),
                    "mc_amount" : row.get("amount"),
                    "raw_date"  : None,
                    "raw_payee" : None,
                    "raw_amount": None,
                    "status"    : "unmatched",
                },
            )
    
    if unmatched_rows:
        unmatched_df = pd.DataFrame(unmatched_rows)
        mc_sheet_df = pd.concat(
            [
                mc_matches.dropna(axis=1, how="all"),
                unmatched_df.dropna(axis=1, how="all"),
            ],
            ignore_index=True,
        )
    else:
        mc_sheet_df = mc_matches.copy()
    
    matched_mc = len(matched_keys)
    unmatched_mc = total_mc - matched_mc
    
    return mc_sheet_df, total_mc, unmatched_mc


def style_mastercard_sheet(workbook) -> None:
    """
    Apply alternating fills to groups of rows in the 'מסטרקארד' sheet,
    grouping by identical 'mc_amount' values.
    """
    mc_sheet_name = "מסטרקארד"
    if mc_sheet_name not in workbook.sheetnames:
        return
    
    ws_mc = workbook[mc_sheet_name]
    
    # Find the column index for 'mc_amount'
    mc_amount_col_idx: Optional[int] = None
    for col_idx in range(1, ws_mc.max_column + 1):
        hdr_val = ws_mc.cell(row=1, column=col_idx).value
        if hdr_val == "mc_amount":
            mc_amount_col_idx = col_idx
            break
    
    if mc_amount_col_idx is None:
        return
    
    amount_to_fill = {}
    fill_idx = 0
    fills = list(MASTERCARD_ROW_FILLS)
    
    for row_idx in range(2, ws_mc.max_row + 1):
        cell = ws_mc.cell(row=row_idx, column=mc_amount_col_idx)
        val = cell.value
        if val is None:
            continue
        
        key = float(val) if isinstance(val, (int, float)) else str(val)
        
        if key not in amount_to_fill:
            amount_to_fill[key] = fills[fill_idx % len(fills)]
            fill_idx += 1
        
        row_fill = amount_to_fill[key]
        
        for col_idx in range(1, ws_mc.max_column + 1):
            ws_mc.cell(row=row_idx, column=col_idx).fill = row_fill


# ---------------------------------------------------------------------------
# Per-month processing
# ---------------------------------------------------------------------------

def process_month(
    year: int,
    month: int,
    filtered_df: pd.DataFrame,
    blocked_df: pd.DataFrame,
    rules_df: pd.DataFrame,
    lookup_df: pd.DataFrame,
) -> None:
    """
    Process a single (year, month) slice of filtered + blocked transactions:
      - normalize & split into Parsed / Income sheets
      - match Mastercard transactions against 9846 raw data
      - write Excel + summary sheets
      - run merged-category refinement
    """
    mm = f"{month:02d}"
    yy = f"20{year % 100:02d}"
    
    # Filter this month's "filtered" rows
    date_dt = pd.to_datetime(filtered_df["date"], format="%d.%m.%y", errors="coerce")
    month_mask = (date_dt.dt.year == year) & (date_dt.dt.month == month)
    df_month = filtered_df.loc[month_mask & date_dt.notna()].copy()
    if df_month.empty:
        return
    
    # Blocked rows for this month (cards, etc.)
    blocked_dates = pd.to_datetime(
        blocked_df["date"],
        format="%d.%m.%y",
        errors="coerce",
    )
    blocked_month = blocked_df[
        (blocked_dates.dt.year == year) & (blocked_dates.dt.month == month)
        ].copy()
    
    # Mastercard subset
    mastercard_mask = (
            blocked_month["payee"].astype(str).str.contains(MASTERCARD_IDENTIFIERS[0], na=False)
            | blocked_month["payee"].astype(str).str.contains(MASTERCARD_IDENTIFIERS[1], na=False)
    )
    mastercard_month = blocked_month[mastercard_mask].copy()
    
    mc_sheet_df, total_mc, unmatched_mc = create_mastercard_report(
        mastercard_month,
        mm=mm,
        yy=yy,
    )
    
    # Normalize and split into Parsed / Income
    output_path = DATA_PROCESSED / f"9016_{mm}_{yy}.xlsx"
    df_parsed, df_income = process_transactions(df=df_month)
    
    df_parsed = df_parsed.dropna(subset=["amount"])
    df_parsed.sort_values("amount", inplace=True)
    
    # Apply payee normalization + category lookup to Parsed sheet
    # Add columns expected by transaction-specific payee rules (safe no-ops if unused).
    if not df_parsed.empty and "payee" in df_parsed.columns:
        if "account" not in df_parsed.columns:
            df_parsed["account"] = "9016"
        # Use absolute value for matching against rule "expense" amounts (rules are typically positive)
        df_parsed["expense"] = pd.to_numeric(df_parsed.get("amount"), errors="coerce").abs()
        df_parsed = apply_payee_rules_and_categories(
            df_parsed,
            rules_df=rules_df,
            lookup_df=lookup_df,
            payee_col="payee",
            account_col="account",
            date_col="date",
            expense_col="expense",
        )
    
    # Do the same for Income sheet if it has a payee column
    if df_income is not None and not df_income.empty and "payee" in df_income.columns:
        df_income = apply_payee_rules_and_categories(
            df_income,
            rules_df=rules_df,
            lookup_df=lookup_df,
            payee_col="payee",
        )
    
    # Simple debug print for date range
    dates = pd.to_datetime(df_parsed["date"], format="%d.%m.%y", errors="coerce").dropna()
    if not dates.empty:
        print(
            f"{mm}.{yy} range:",
            dates.min().strftime("%Y-%m"),
            dates.max().strftime("%Y-%m"),
        )
    
    # ---- Write Excel ----
    DATA_PROCESSED.mkdir(parents=True, exist_ok=True)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_parsed.to_excel(writer, sheet_name="Parsed", index=False)
        df_income.to_excel(writer, sheet_name="Income", index=False)
        
        summary_df = pd.DataFrame(
            [
                {
                    "month"                    : f"{mm}.{yy}",
                    "total_mastercard_rows"    : total_mc,
                    "matched_mastercard_rows"  : total_mc - unmatched_mc,
                    "unmatched_mastercard_rows": unmatched_mc,
                },
            ],
        )
        summary_df.to_excel(writer, sheet_name="MC_Summary", index=False)
        
        # Per-card sheets
        for kw in CARD_SUMMARY_SHEET_KEYWORDS:
            kw_month_df = blocked_month[
                blocked_month["payee"].astype(str).str.contains(kw, na=False)
            ].copy()
            
            if kw == "מסטרקארד":
                sheet_df = mc_sheet_df
            else:
                sheet_df = kw_month_df
            
            if sheet_df is not None and not sheet_df.empty:
                sheet_name = kw[:28]
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        workbook = writer.book
        
        # Generic formatting: amounts, column widths, filters, freeze panes, etc.
        format_workbook_default(workbook)
        
        # Color the Mastercard sheet nicely
        style_mastercard_sheet(workbook)
    
    print(f"Done. Wrote {len(df_month)} rows to: {output_path.resolve()}")
    refine_merged_categories(mm, yy)


# ---------------------------------------------------------------------------
# main()
# ---------------------------------------------------------------------------

def main() -> None:
    # 1. Load the main 9016 XLS from the bank
    input_path = TNUOT_9016
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")
    
    parsed_df = load_and_clean_9016_xls(input_path)
    if parsed_df.empty:
        raise SystemExit("No transactions found in 9016 XLS file")
    
    parsed_df["amount"] = pd.to_numeric(parsed_df["amount"], errors="coerce").round(2)
    
    # 2. Load payee normalization + lookup tables
    rules_df, lookup_df = load_payee_resources()
    
    # 3. Split into "filtered" vs "blocked" (credit cards, etc.)
    blocked_keywords = load_blocked_payee_keywords()
    pattern = "|".join(re.escape(kw) for kw in blocked_keywords)
    
    mask = parsed_df["payee"].astype(str).str.contains(pattern, na=False)
    filtered_df = parsed_df[~mask].copy()
    blocked_df = parsed_df[mask].copy()
    
    # 4. Determine which (year, month) combinations exist and process each
    date_dt = pd.to_datetime(filtered_df["date"], format="%d.%m.%y", errors="coerce")
    if date_dt.notna().sum() == 0:
        raise SystemExit("No valid dates found; cannot determine mm_yy.")
    
    valid_mask = date_dt.notna()
    filtered_df = filtered_df.loc[valid_mask].copy()
    date_dt = date_dt[valid_mask]
    
    months_years = sorted(
        set(zip(date_dt.dt.year.astype(int), date_dt.dt.month.astype(int))),
    )
    
    for year, month in months_years:
        process_month(
            year=year,
            month=month,
            filtered_df=filtered_df,
            blocked_df=blocked_df,
            rules_df=rules_df,
            lookup_df=lookup_df,
        )


if __name__ == "__main__":
    main()
